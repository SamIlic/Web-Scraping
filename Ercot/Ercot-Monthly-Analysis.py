# -*- coding: utf-8 -*-
import scrapy # needed to scrape
import xlrd   # used to easily import xlsx file 
import json
import re
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import datetime
from datetime import timedelta

##### NOTE
# PART 1: This script writes all monthly Ercot data to its own tab without affecting the "Master Data" tab 
# PART 2: This script will perform analysis on the monthly data 


### PART 1 - Clean Data
##########################################################################################
##########################################################################################
##########################################################################################
file_path = r"/Users/YoungFreeesh/Visual Studio Code/_Python/Web Scraping/Ercot/MASTER-Ercot.xlsx"
df = pd.read_excel(file_path, sheet_name = 'Master Data') # read all data from "Master Data" tab in the "MASTER-Ercot" workbook
headers = list(df.columns.values) # get the headers of "Master Data"
df = pd.DataFrame(df) # convert df to a Date Frame
print('~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~')

### Get all Unique Months in the data frame
dateArray = np.array(df.iloc[:, 0]) # convert dates column in df to a numpy array
monthsArray = [] # initialize array
for x in range(dateArray.shape[0]): # change format of dates: 05/10/2018 --> 05-2018
    temp = dateArray[x]
    #print(str(temp[:2]) + '-' + str(temp[6:]))
    #print(str(temp))
    #monthsArrayUnique = np.unique(monthsArray) # Unique Months, hence get all unique strings of the form: 'mm-yyyy'
    #print("Unique Months: ", monthsArrayUnique)
    # if (str(temp[:2]) + '-' + str(temp[6:]) == "08-2018"):
    #     monthsArray.append("08-2018")
    monthsArray.append(str(temp[:2]) + '-' + str(temp[6:]))
    #print(monthsArray)
monthsArrayUnique = np.unique(monthsArray) # Unique Months, hence get all unique strings of the form: 'mm-yyyy'
print("Unique Months: ", monthsArrayUnique)


##### Use PANDAS to write to an Excel file and create tabs
##########################################################################################
#file_path_HardDrive = r"/Users/YoungFreeesh/Visual Studio Code/_Python/Web Scraping/Ercot/Test-Ercot-Scrape.xlsx"
#file_path_Dropbox   = r"/Users/YoungFreeesh/Dropbox/Ercot Data/Test-Ercot-Scrape.xlsx"
file_path_HardDrive = r"/Users/YoungFreeesh/Visual Studio Code/_Python/Web Scraping/Ercot/MASTER-Ercot.xlsx"
file_path_Dropbox   = r"/Users/YoungFreeesh/Dropbox/Ercot Data/MASTER-Ercot.xlsx"

### For Ercot Summary Page - Calculations (LZ_SOUTH)
# read all data from "Master Data" tab from "MASTER-Ercot" 
dfMASTER = pd.read_excel(file_path_HardDrive, sheet_name = 'Master Data')

writer_HardDrive = pd.ExcelWriter(file_path_HardDrive, engine='openpyxl')
writer_Dropbox   = pd.ExcelWriter(file_path_Dropbox  , engine='openpyxl')

book_HardDrive = load_workbook(file_path_HardDrive)
book_Dropbox   = load_workbook(file_path_Dropbox)

writer_HardDrive.book = book_HardDrive
writer_Dropbox.book   = book_Dropbox

writer_HardDrive.sheets = dict((ws.title, ws) for ws in book_HardDrive.worksheets)
writer_Dropbox.sheets   = dict((ws.title, ws) for ws in book_Dropbox.worksheets)


### Create a unique Excel Worksheet/tab for each month of data in the Master Date tab
# This tab will contain all the price data for that particular month
# This loop will create a tab for the month if it doesn't alreay exist
# This loop will overwrite any data already in the months tab in columns A-P
# This loop will not affect and data or formulas or graphs Beyond column Q
# This loop will not affect any other tabs 
monthsArray = np.array(monthsArray) # convert monthsArray to a numpy array
for month in monthsArrayUnique:
    print("Tab Created: ", month)
    indices = [i for i, x in enumerate(monthsArray) if x == month] # get indices for month
    monthDF = pd.DataFrame(data = np.array(df.iloc[indices, :]), columns = headers)  # create data frame
    
    monthDF.to_excel(writer_HardDrive, startrow= 0 , index=False, sheet_name=str(month)) # write to "MASTER-Ercot.xlsx" spreadsheet
    monthDF.to_excel(writer_Dropbox  , startrow= 0 , index=False, sheet_name=str(month)) # write to "MASTER-Ercot.xlsx" spreadsheet

print('~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~')


# End of PART 1
##########################################################################################




### PART 2 - Analyze Data
##########################################################################################
##########################################################################################
##########################################################################################

###Ercot Summary Page - Calculations (LZ_SOUTH)

### Refine the DataFrame
#Only Take LZ_SOUTH
dfMASTER_LZ_SOUTH =  dfMASTER[['Oper Day', 'Interval Ending', 'LZ_SOUTH']].copy(deep=True)
dfMASTER_LZ_SOUTH['Oper Day'] = pd.to_datetime(dfMASTER_LZ_SOUTH['Oper Day'])

# Reset index to Oper Day
dfMASTER_LZ_SOUTH = dfMASTER_LZ_SOUTH.set_index('Oper Day')

##### Begin Calculations #####
# 1) Avg. Daily Price
daily_avg = pd.DataFrame() # Initialize dataframe
daily_avg['Mean'] = dfMASTER_LZ_SOUTH.LZ_SOUTH.resample('D').mean() # Resample df & compute mean

# 2) Avg. Price - Power was cut off from 3:15pm to 4:15pm 
#       Remove Prices from: 1515 to 1615 --> (1530, 1545, 1600, 1615)
dfMASTER_LZ_SOUTH_Optimized = dfMASTER_LZ_SOUTH[ ( (dfMASTER_LZ_SOUTH['Interval Ending'] <= 1515) | (dfMASTER_LZ_SOUTH['Interval Ending'] >= 1630) ) ].copy(deep=True)
daily_avg_Optimized = pd.DataFrame()
daily_avg_Optimized['Mean'] = dfMASTER_LZ_SOUTH_Optimized.LZ_SOUTH.resample('D').mean() # Resample df & compute mean

# 3)Difference between (1) and (2) 
difference = daily_avg - daily_avg_Optimized

### Create Daily Summary Worksheet
# Create Summary DataFrame
dfSummary = pd.DataFrame()
dfSummary['Avg Daily Price'] = daily_avg['Mean']
dfSummary['Avg Daily Price - Optimized'] = daily_avg_Optimized['Mean']
dfSummary['Difference'] = difference['Mean']
# Write to worksheet
dfSummary.to_excel(writer_HardDrive, startrow= 0 , index=True, sheet_name= 'Daily Summary') # write to "MASTER-Ercot.xlsx" spreadsheet
dfSummary.to_excel(writer_Dropbox  , startrow= 0 , index=True, sheet_name= 'Daily Summary') # write to "MASTER-Ercot.xlsx" spreadsheet

print('~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~')
print('Tab Created: Daily Summary')
print('~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~')
# End of PART 2 
##########################################################################################




### Close the Pandas Excel writer and output the Excel file.
writer_HardDrive.save()
writer_Dropbox.save()

writer_HardDrive.close()
writer_Dropbox.close()

































