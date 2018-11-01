# -*- coding: utf-8 -*-
import scrapy # needed to scrape
import xlrd   # used to easily import xlsx file 
import json
import re
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import datetime
from datetime import timedelta

class StoreDataCurrent(scrapy.Spider):
    name = 'Ercotbot'

    ### Get Dates: Today, Yesterday, Last Day in "MASTER-Ercot"
    dateToday = str(datetime.datetime.today().strftime('%Y%m%d')) # Today, as an Integer
    dateYesterday = str(datetime.date.fromordinal(datetime.date.today().toordinal()-1)).replace('-', '')
    
    ### Create dates for URL
    df = pd.read_excel('MASTER-Ercot.xlsx', sheet_name = 'Master Data')
    df = pd.DataFrame(df)
    lastDate = str(df.iat[df.shape[0] - 1, 0]) # get the last scraped date in "MASTER-Ercot"
    splitDate = lastDate.split('/') # split up the date and get rid of "/" 
    SD0 = splitDate[0];  SD1 = splitDate[1];  SD2 = splitDate[2]
    splitDate[0] = SD2;  splitDate[1] = SD0;  splitDate[2] = SD1
    lastDate = str(''.join(splitDate))
    # lastDate = str(20180810)
    print('dateToday: ', dateToday)
    print('dateYesterday: ', dateYesterday)
    print('lastDate: ', lastDate)
    print('Appened? ', int(lastDate) != int(dateYesterday))  

    
    allowed_domains = ['ercot.com']
    start_urls      = ['http://ercot.com/content/cdr/html/{}_real_time_spp'.format(dateYesterday)]
    print("``````````````````````````````````````````````````````````````````````````````")


################################################################################################
################################################################################################

    ###
    # Scrape Daily Ercot data and Append to "MASTER-Ercot" file
    ###
    def parse(self, response):
        self.logger.info('A response has arrived from %s just arroved form ', response.url)
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        
        ### Scrape table Headers and Values (energy prices)
        #headers = response.css(".headerValueClass::text").extract()
        values  = response.css(".labelClassCenter::text").extract()


        ### convert to a data frame and append Header  
        values = np.array(values) # turn ercot data into an numpy array
        values = np.reshape(values, (int(len(values)/16), 16)) #reshape the array to be in the same format as it is online
        #frameHeaders = pd.DataFrame(data = values[0:][0:], columns = headers) # --> makes it easier to write to xlsx file 
        frame = pd.DataFrame(data = values[1:][0:], columns = values[0][0:]) # --> makes it easier to write to xlsx file 
        
        
        ### Get Dates: Today, Yesterday, Last Day in "MASTER-Ercot"
        dateToday = str(datetime.datetime.today().strftime('%Y%m%d')) # Today, as an Integer
        dateYesterday = str(datetime.date.fromordinal(datetime.date.today().toordinal()-1)).replace('-', '')
        
        df = pd.read_excel('MASTER-Ercot.xlsx', sheet_name = 'Master Data')
        df = pd.DataFrame(df)
        lastDate = str(df.iat[df.shape[0] - 1, 0]) # get the last scraped date in "MASTER-Ercot"
        splitDate = lastDate.split('/') # split up the date and get rid of "/" 
        SD0 = splitDate[0];  SD1 = splitDate[1];  SD2 = splitDate[2]
        splitDate[0] = SD2;  splitDate[1] = SD0;  splitDate[2] = SD1
        lastDate = str(''.join(splitDate))
        # lastDate = str(20180810)

        print('dateToday: ', dateToday)
        print('dateYesterday: ', dateYesterday)
        print('lastDate: ', lastDate)
        print('Appened? ', int(lastDate) != int(dateYesterday))  
      

        ### Append new data to the "MASTER-Ercot" spreadsheet
        # If we already appended yesterdays data --> do not append again
        if ( int(lastDate) != int(dateYesterday) ): # this prevents us from writing the same data to the spreadsheet multiple times
            
            # Write to Current Working Directory
            writer = pd.ExcelWriter('MASTER-Ercot.xlsx', engine='openpyxl') 
            book = load_workbook('MASTER-Ercot.xlsx')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            frame.to_excel(writer, startrow=len(df)+1 , index=False, sheet_name = 'Master Data')
            writer.save()
            writer.close()

            # Write to Dropbox
            out_path = r"/Users/YoungFreeesh/Dropbox/Ercot Data/MASTER-Ercot.xlsx" # the `r` prefix means raw string
            writer = pd.ExcelWriter(out_path, engine='openpyxl') 
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            frame.to_excel(writer, startrow=len(df)+1 , index=False, sheet_name = 'Master Data')
            writer.save()
            writer.close()
            
            
            
### Other Pandas stuff
#################################################################################################
#################################################################################################
#################################################################################################
#################################################################################################
#################################################################################################

        # import subprocess
        # import os
        # path = r"/Users/YoungFreeesh/Visual\ Studio\ Code/_Python/Web\ Scraping"
        # os.chdir(path) # set current working directory
        # subprocess.call("Ercot-Monthly-Analysis.py", shell=True)


        ### Pandas to Open file and print contentents
        #################################################################################################
        # my_sheet_name = 'Sheet1' 
        # df = pd.read_excel('pandas_simple.xlsx', sheet_name = my_sheet_name)
        # print(df.head()) # shows headers with top 5 rows 

        ### Pandas to Append to file
        #################################################################################################
        # df2 = pd.DataFrame({'Data': [50000, 200, 300, 200, 150, 300, 450], 
        #                     'Name1': 'name1',
        #                     'Name2': 'name2',
        #                     'Name3': 'name3',
        #                     'Name4': 'name4'
        # })
        
        #path = r"C:\Users\fedel\Desktop\excelData\PhD_data.xlsx"
        # book = load_workbook('pandas_simple.xlsx')
        # writer = pd.ExcelWriter('pandas_simple.xlsx', engine='openpyxl') 

        # df = pd.read_excel('MASTER-Ercot.xlsx', sheet_name = 'Sheet1')
        # book = load_workbook('MASTER-Ercot.xlsx')
        # writer = pd.ExcelWriter('MASTER-Ercot.xlsx', engine='openpyxl') 
        # writer.book = book
        # writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        # #frame.to_excel(writer, startrow=len(df)+1 , index=False, sheet_name = 'Sheet1')

        # writer.save()
        # writer.close()



        #################################################################################################
        ### Pandas to write a file
        #################################################################################################
        #################################################################################################
        # Use PANDAS to write a file
        # # Create a Pandas dataframe from the data.
        # #df = pd.DataFrame({'Data': [10, 20, 30, 20, 15, 30, 45]})

        # # Create a Pandas Excel writer using XlsxWriter as the engine.
        # writer = pd.ExcelWriter('pandas_simple2.xlsx', engine='xlsxwriter')

        # # Convert the dataframe to an XlsxWriter Excel object.
        # #df.to_excel(writer, sheet_name='Sheet1')
        # frame.to_excel(writer, sheet_name='Sheet1')

        # # Close the Pandas Excel writer and output the Excel file.
        # writer.save()
        # writer.close()




        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")






